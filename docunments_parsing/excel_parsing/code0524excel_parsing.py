# !/usr/bin/python3
# -*- coding: utf-8 -*-
"""
@Author         :  黄英杰
@Version        :  0.0.1
------------------------------------
@File           :  code0524excel_parsing.py
@Description    :  
@CreateTime     :  2020/5/24 0024 下午 20:16
------------------------------------
@ModifyTime     :  
"""


from openpyxl import load_workbook

class ExcelParsing:
    def __init__(self,filepath, sheet_names=None):
        self._filepath = filepath
        self.resutl = {}
        self.wb = load_workbook(self._filepath)
        self.workbook = self.wb.active
        if sheet_names:
            if sheet_names == "all":
                sheet_names = self.wb.sheetnames
            for sheet_name in sheet_names:
                self.workbook = self.wb[sheet_name]
                self.resutl[sheet_name]=self.get_cols()

    def get_cols(self,min_col=None, max_col=None, only=True):
        cols = []
        for col in self.workbook.iter_cols(min_col=min_col, max_col=max_col, values_only=only):
            col_cells = []
            for cell in col:
                col_cells.append(cell)
            cols.append(col_cells)
        return cols

    def get_rows(self,min_row=None, max_row=None, only=True):
        rows = []
        for row in self.workbook.iter_rows(min_row=min_row, max_row=max_row, values_only=only):
            row_cells = []
            for cell in row:
                row_cells.append(cell)
            rows.append(row_cells)
        return rows

    def save_files(self):
        pass

if __name__ == '__main__':
    excel = ExcelParsing("")
    cols=excel.get_cols()
    print(cols)
    print(excel.resutl)
    cols=excel.get_cols()
    print(cols)
