# !/usr/bin/python3
# -*- coding: utf-8 -*-
"""
@Author         :  黄英杰
@Version        :  1.0.0
------------------------------------
@File           :  code20210224_write__execl.py
@Description    :  
@CreateTime     :  2021/2/24 0024 下午 19:45
@微信号          :  HelperRobot
------------------------------------
@ModifyTime     :  
"""
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class ExcelWrite:
    def __init__(self,filepath=None, save_name=None, sheet_name=None):
        self._filepath = filepath
        self._sheet_name = sheet_name
        self._save_name = save_name

    def write_rows(self, row_datas):
        self.wb.append(row_datas)

    def write_columns(self, column_datas=None, column_idx=None):
        for row, data in enumerate(column_datas, 1):
            self.wb.cell(value=data, row=row, column=column_idx if column_idx else 1)

    def __enter__(self):
        self.workbook = load_workbook(self._filepath)[self._sheet_name] if self._filepath else Workbook()
        self.wb = self.workbook[self._sheet_name] if self._filepath else self.workbook.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._sheet_name:
            self.wb.title = self._sheet_name
        self.workbook.save(self._save_name)
        self.workbook.close()

if __name__ == '__main__':
    with ExcelWrite(save_name="test.xlsx", sheet_name="666") as excel_obj:
        excel_obj.write_columns(range(5))
